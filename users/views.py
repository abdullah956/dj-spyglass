from .forms import CustomUserCreationForm
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
import pyotp
from django.contrib.auth.forms import SetPasswordForm
from properties.models import Property
from .models import Agent, Homeowner, Assistant, User , Contact , NewsletterSubscription
from .forms import CustomUserCreationForm , UserEditForm
from openpyxl import Workbook
from django.http import HttpResponse
from subscriptions.models import Subscription
import pytz

# for home 
# def home_view(request):
#     properties = Property.objects.filter(approval_status=True)    
#     return render(request, 'home.html', {'properties': properties})
def home_view(request):
    properties = Property.objects.filter(approval_status=True)
    active_subscription = None
    if request.user.is_authenticated:
        active_subscription = request.user.subscription_set.filter(is_active=True, payment_successful=True).exists()
    print(active_subscription)
    return render(request, 'home.html', {'properties': properties, 'active_subscription': active_subscription})


# to register
# def register_view(request):
#     if request.method == 'POST':
#         form = CustomUserCreationForm(request.POST)
#         if form.is_valid():
#             user = form.save(commit=False)
#             user.role = 'Agent'
#             user.save()
#             Agent.objects.create(user=user)

#             otp = pyotp.TOTP(settings.OTP_SECRET_KEY)
#             otp_code = otp.now()
#             send_mail(
#                 'Your OTP Code',
#                 f'Your OTP code is {otp_code}',
#                 settings.EMAIL_HOST_USER,
#                 [user.email],
#                 fail_silently=False,
#             )

#             request.session['otp_email'] = user.email
#             request.session['otp_code'] = otp_code
#             messages.success(request, 'Account created successfully! Please verify your email with the OTP sent.')
#             return redirect('verify_otp')
#         else:
#             messages.error(request, 'There was an error in the registration form. Please correct the issues and try again.')
#     else:
#         form = CustomUserCreationForm()

#     return render(request, 'users/register.html', {'form': form})

def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'Agent'
            user.save()
            Agent.objects.create(user=user)

            otp = pyotp.TOTP(settings.OTP_SECRET_KEY)
            otp_code = otp.now()
            send_mail(
                'Your OTP Code',
                f'Your OTP code is {otp_code}',
                settings.EMAIL_HOST_USER,
                [user.email],
                fail_silently=False,
            )

            request.session['otp_email'] = user.email
            request.session['otp_code'] = otp_code
            messages.success(request, 'Account created successfully! Please verify your email with the OTP sent.')
            return redirect('verify_otp')
        else:
            messages.error(request, 'There was an error in the registration form. Please correct the issues and try again.')
    else:
        form = CustomUserCreationForm()

    return render(request, 'users/register.html', {'form': form})

# to login 
def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            if user.is_verified:
                login(request, user)
                messages.success(request, 'You have successfully logged in.')
                return redirect('home')
            else:
                otp = pyotp.TOTP(settings.OTP_SECRET_KEY)
                otp_code = otp.now()
                send_mail(
                    'Your OTP Code',
                    f'Your OTP code is {otp_code}',
                    settings.EMAIL_HOST_USER,
                    [user.email],
                    fail_silently=False,
                )
                request.session['otp_email'] = user.email
                request.session['otp_code'] = otp_code
                messages.info(request, 'Your account is not verified. Please enter the OTP sent to your email.')
                return redirect('verify_otp')
        else:
            messages.error(request, 'Invalid email or password. Please try again.')
            return redirect('login')

    return render(request, 'users/login.html')
# to logout
def logout_view(request):
    logout(request)
    return redirect('login')

# to verify login otp
def verify_otp(request):
    if request.method == 'POST':
        user_otp_code = request.POST.get('otp_code')
        session_otp_code = request.session.get('otp_code')
        email = request.session.get('otp_email')

        if user_otp_code == session_otp_code:
            try:
                user = User.objects.get(email=email)
                user.is_verified = True
                user.save()
                login(request, user)
                messages.success(request, 'Your account has been successfully verified and you are now logged in.')
                return redirect('home')
            except User.DoesNotExist:
                messages.error(request, 'User not found. Please try again.')
                return redirect('verify_otp')
        else:
            messages.error(request, 'Invalid OTP. Please try again.')
            return redirect('verify_otp')

    return render(request, 'users/verify_otp.html')

# to request password reset
def password_reset_request(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        
        try:
            user = User.objects.get(email=email)
            otp = pyotp.TOTP(settings.OTP_SECRET_KEY)
            otp_code = otp.now()

            send_mail(
                'Your OTP Code for Password Reset',
                f'Your OTP code is {otp_code}',
                settings.EMAIL_HOST_USER,
                [user.email],
                fail_silently=False,
            )

            request.session['otp_email'] = user.email
            request.session['otp_code'] = otp_code

            messages.success(request, 'An OTP has been sent to your email for password reset.')
            return redirect('verify_otp_for_password_reset')

        except User.DoesNotExist:
            messages.error(request, 'The email you entered does not exist.')
            return redirect('password_reset_request')

    return render(request, 'users/password_reset_request.html')

# to verify password resert otp
def verify_otp_for_password_reset(request):
    if request.method == 'POST':
        user_otp_code = request.POST.get('otp_code')
        session_otp_code = request.session.get('otp_code')
        email = request.session.get('otp_email')

        if user_otp_code == session_otp_code:
            try:
                user = get_user_model().objects.get(email=email)
                messages.success(request, 'OTP verified successfully. You can now change your password.')
                return redirect('password_change', user_id=user.id)
            except get_user_model().DoesNotExist:
                messages.error(request, 'User not found. Please try again.')
                return redirect('verify_otp_for_password_reset')
        else:
            messages.error(request, 'Invalid OTP. Please try again.')
            return redirect('verify_otp_for_password_reset')

    return render(request, 'users/verify_otp_for_password_reset.html')

# to change password
def password_change(request, user_id):
    user = get_user_model().objects.get(id=user_id)
    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)
        if form.is_valid():
            form.save()
            user = authenticate(request, username=user.email, password=request.POST.get('new_password1'))
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                messages.error(request, 'Authentication failed. Please try logging in again.')
    else:
        form = SetPasswordForm(user)
    return render(request, 'users/password_change.html', {'form': form})

# for contact page 
def contact_view(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        address = request.POST.get('address')
        Contact.objects.create(name=name, email=email, address=address)
        messages.success(request, 'Your message has been sent successfully!')
        return redirect('contact')
    return render(request, 'contact/contact.html')

# for newsletter
from django.db import IntegrityError
def newsletter_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        
        # Check if the email field is empty
        if not email:
            messages.error(request, 'Please enter a valid email address.')
            return redirect('home')

        # Check if the email already exists
        if NewsletterSubscription.objects.filter(email=email).exists():
            messages.warning(request, 'You are already subscribed!')
        else:
            try:
                # Attempt to create the subscription
                NewsletterSubscription.objects.create(email=email)
                messages.success(request, 'Thank you for subscribing to our newsletter!')
            except IntegrityError:
                messages.error(request, 'There was an error with your subscription. Please try again.')

        return redirect('home')
    
    return render(request, 'home.html')

#  edit user profile
def edit_profile(request):
    user = get_object_or_404(User, email=request.user.email)

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "User information updated successfully.")
            return redirect('edit_profile')
    else:
        form = UserEditForm(instance=user)

    return render(request, 'users/profile_edit.html', {'form': form})

# admin dashboard
def admin_dashboard(request):
    total_agents = Agent.objects.count()
    total_assistants = Assistant.objects.count()
    total_homeowners = Homeowner.objects.count()
    total_properties = Property.objects.count()
    total_subscribers = NewsletterSubscription.objects.count()

    context = {
        'total_agents': total_agents,
        'total_assistants': total_assistants,
        'total_homeowners': total_homeowners,
        'total_properties': total_properties,
        'total_subscribers': total_subscribers,
    }
    
    return render(request, 'users/dashboard_landingpage.html', context)

# subscriber lists
def subscribers_list(request):
    subscribers = NewsletterSubscription.objects.all()
    context = {
        'subscribers': subscribers,
    }
    return render(request, 'users/subscribers_list.html', context)

# download subcribers
def download_subscribers(request):
    subscribers = NewsletterSubscription.objects.all()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Subscribers'
    
    worksheet.append(['Email'])
    for subscriber in subscribers:
        worksheet.append([subscriber.email])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=subscribers.xlsx'
    
    workbook.save(response)
    return response

# contacts 
def contact_list(request):
    contacts = Contact.objects.all()
    return render(request, 'users/contact_list.html', {'contacts': contacts})

# contact download
def download_contacts_excel(request):
    contacts = Contact.objects.all()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Contacts'
    headers = ['Name', 'Email', 'Message']
    worksheet.append(headers)
    for contact in contacts:
        worksheet.append([contact.name, contact.email, contact.address])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="contacts.xlsx"'
    workbook.save(response)
    return response


# all subscriptions view
def all_subscriptions_view(request):
    subscriptions = Subscription.objects.all()
    return render(request, 'users/all_subscriptions.html', {'subscriptions': subscriptions})

#download subscription
def download_subscriptions_excel(request):
    subscriptions = Subscription.objects.all()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Subscriptions'
    headers = ['User', 'Subscription Type', 'Active', 'Amount', 'Start Date', 'End Date', 'Payment Successful']
    worksheet.append(headers)

    for subscription in subscriptions:
        start_date = subscription.start_date.astimezone(pytz.UTC).isoformat()
        end_date = subscription.end_date.astimezone(pytz.UTC).isoformat()
        worksheet.append([
            subscription.user.email,
            subscription.subscription_type,
            subscription.is_active,
            subscription.amount,
            start_date,
            end_date,
            subscription.payment_successful
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="subscriptions.xlsx"'
    workbook.save(response)
    return response